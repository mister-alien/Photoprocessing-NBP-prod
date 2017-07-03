from openpyxl import Workbook
from openpyxl import load_workbook
import os, shutil

#######################################
# Correct config for test data:
# 
# offset = 2
# prefix='IMG0000'
# namecol='A'
# txcol='D'
# platecol='E'                               
#
# This file is only function
# definitions however, so run the gui
# script for best results.
######################################
def renstuff( spreadsheet, directory, namecol, platecol, txcol, prefix, offset ):
   
    Tri=None
   

    wb = load_workbook(spreadsheet)
    ws=wb.active                                # Assuming the format of only one worksheet per workbook, if this changes we will
                                                # have to fix this
    n= ws.max_row                               # max_row to find out how much we will have to iterate... if there are empty rows at the end that is ok

    errs =["" for  x in range(2*n)]
    eind=0
    
    dest =directory+'\\..\\processed\\'         # the destination directory in full for future use, used ~8 times 
    
    if not os.path.exists(dest):                # If the destination directory doesn't exist, make it
        os.makedirs(dest)
    else:
        shutil.rmtree(dest)                     # If it does exist, destroy its contents... why? If prior attempts at naming went astray, 
        os.makedirs(dest)                       # there can be junk poorly named files that are best deleted.

        
    for x in range(1,n-offset):
        name=ws[namecol +str(offset+x)]         # cell containing the tx name
        platenum=ws[platecol +str(offset+x)]    # current cell containing plate pic number
        txnum=ws[txcol +str(offset+x)]          # current cell containing tx pic number
        
        # print(name.value)                       # Which item is currently being processed (2 photo per item)

        ### All of the ugly directory work is here, 
        photop=directory+'\\'+prefix[0:(len(prefix)-len(str(platenum.value)))]+str(platenum.value)+'.jpg'   # Plate filename+path(RAW)
        photot=directory+'\\'+prefix[0:(len(prefix)-len(str(txnum.value)))]+str(txnum.value)+'.jpg'         # Tx filename+path(RAW)
        
        triname=dest+name.value+'.jpg'          # dest triphase filename+path
        txname=dest+name.value+' T.jpg'         # dest tx pic filename+path
        platename=dest+name.value+' P.jpg'      # dest plate (single phase) filename+path
        ###

        # Herein lies the actual copying
        if (platenum.value) and (txnum.value):  # If it's a SINGLE PHASE transformer
            try:
                shutil.copy(photop,platename)   # Copy both tx and plate pic over to the destination
            except IOError, e:
                errs[eind] = photop+" Error: "+str(e)+"\n"
                eind =eind+1
            try:
                shutil.copy(photot,txname)
            except IOError, e:
                errs[eind] = photot+" Error: "+str(e)+"\n"
                eind =eind+1
        elif (txnum.value) and (platenum.value == None):
#            Tri='R'                             # Detecting TRI-PHASE! copies the transformer pic, and sets the threephase flag for the next row 
            try:
                shutil.copy(photot,txname)      # Script assumes R-C-F order, if this is not the case we DONT check for it, but it should
            except IOError, e:                  # still name correctly in that case.
                errs[eind] = photot+" Error: "+str(e)+"\n"
                eind =eind+1
        elif (txnum.value == None) and platenum.value:
                                                # Still TRI PHASE, plate pic detection.
                                                # Basically depending on the tri flag, we rename  the photos,
                                                # BUT you can see now that no logic actually depends on the tri flag, and this can be commented out.
                                                # if the large block of code is commented out, all is well.
#            if Tri=='R':                        
#                shutil.copy(photop,triname)                                                     
#                Tri = 'C'                       
#            elif Tri=='C':
#                shutil.copy(photop,triname)
#               Tri = 'F'
#            elif Tri=='F':
#                shutil.copy(photop,triname)
#                Tri = None
            try:
                shutil.copy(photop,triname)
            except IOError, e:
                errs[eind] = photop+" Error: "+str(e)+"\n"
                eind =eind+1
                
    return errs
