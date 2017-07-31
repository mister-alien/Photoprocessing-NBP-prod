from openpyxl import Workbook
from openpyxl import load_workbook

### DUPLICATE CHECKING FUNCTION... kinda hacky but works
# Christopher Logan 2017
# 
def check_duplicates(spreadsheet, txcol, platecol, namecol,offset):

    wb = load_workbook(spreadsheet)                     # Standard workbook selection code.
    ws=wb.active                                        # Assume active worksheet.. ( maybe not the best choice but that's okay for now.)
    
    n= ws.max_row                                       # Max row index 
    mat = [None for y in range(n*2)]                    # If EVERY single item is unique and single phase, then we will have 2*n-4 items. so i keep some padding
    dupes =["" for  x in range(n)]                      # Why? So we can have lots of possible duplicates. This is an unreasonably large size for the buffer
                                                        # unless NO QA was done but you gotta code for the worst case.

    dupeind=0                                           # How many duplicates?
    ind= 0                                              # how many unique numbers?

    for z in range(1,n-offset+1):                         # check all rows... woo
        curplate=ws[platecol+str(z+offset)].value       # Get the numbers for tx and plate and the tx name
        curtx=ws[txcol+str(z+offset)].value
        curpole=ws[namecol+str(z+offset)].value

        ## IF IT'S A SINGLE PHASE TX
        if curplate and curtx:                                  
            olddupe=dupeind                             # Store the old value of the number of duplicates for reference
            for zz in range(0, ind):                    # Check if the number for the tx pic occurs anywhere, if so 
                if mat[zz]==curtx:                      # put that Tx name + which picture , then increment the number 
                    dupes[dupeind]=curpole+' tx pic\n'  # of duplicates
                    dupeind+=1
                    break
            if dupeind <= olddupe:                      # If the duplicates didnt increase we can say the number is unique..
                mat[ind]=curtx
                ind+=1                                  # Increase number of unique digits.
                
            olddupe=dupeind                             # Repeat the above logic EXACTLY for the plate picture # 
            for yy in range(0, ind):
                if mat[yy]==curplate:                   # (This is the same with all 3 cases, copy->paste 
                    dupes[dupeind]=curpole+' plate pic\n'
                    dupeind+=1                          # (Which VERY LIKELY means that there is a lot of simplifying to do.)
                    break
            if dupeind <= olddupe:
                mat[ind]=curplate
                ind+=1
                
        ## IF IT'S THE TX PIC OF A THREE PHASE TX (no new logic introduced)
        elif curtx and (curplate == None):
            olddupe=dupeind
            for zz in range(0, ind):
                if mat[zz]==curtx:
                    dupes[dupeind]=curtx+'3phase tx pic\n'
                    dupeind+=1
                    break
            if dupeind <= olddupe:
                mat[ind]=curtx
                ind+=1
                
        ## IF IT'S THE PLATE PIC (ANY OF THEM) OF A THREE PHASE TX (no new logic introduced)
        elif curplate and (curtx == None):
            olddupe=dupeind
            for zz in range(0, ind):
                if mat[zz]==curplate:
                    dupes[dupeind]=curpole+' plate pic\n'
                    dupeind+=1
                    break
            if dupeind <= olddupe:
                mat[ind]=curplate
                ind+=1
                
    ## Return the list of duplicates.
    return dupes 
            
