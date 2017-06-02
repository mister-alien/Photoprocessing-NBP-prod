from openpyxl import Workbook
from openpyxl import load_workbook
import os, shutil

offset = 2

namecol='A'
txcol='D'
platecol='E'
Tri=None
prefix='IMG0000'

wb = load_workbook('Line_1234.xlsx')
ws=wb.active
#a=ws[platecol +str(offset+ind)]
n= ws.max_row

if not os.path.exists('processed'):
    os.makedirs('processed')
for x in range(1,n-offset):
    name=ws[namecol +str(offset+x)]
    platenum=ws[platecol +str(offset+x)]
    txnum=ws[txcol +str(offset+x)]
    print(name.value)
    if (platenum.value) and (txnum.value):
        shutil.copyfile('raw\\'+prefix[0:(len(prefix)-len(str(platenum.value)))]+str(platenum.value)+'.jpg','processed\\'+name.value+' P.jpg')
#        os.rename('raw\\'+prefix[0:(len(prefix)-len(str(platenum.value)))]+str(platenum.value)+'.jpg',name.value+' P.jpg')
        shutil.copyfile('raw\\'+prefix[0:(len(prefix)-len(str(txnum.value)))]+str(txnum.value)+'.jpg','processed\\'+name.value+' T.jpg')

#        os.rename('raw\\'+prefix[0:(len(prefix)-len(str(txnum.value)))]+str(txnum.value)+'.jpg',name.value+' T.jpg')
    elif (txnum.value) and (platenum.value == None):
        Tri='R'
        shutil.copyfile('raw\\'+prefix[0:(len(prefix)-len(str(txnum.value)))]+str(txnum.value)+'.jpg','processed\\'+name.value+' T.jpg')

#        os.rename('raw\\'+prefix[0:(len(prefix)-len(str(txnum.value)))]+str(txnum.value)+'.jpg',name.value+' T.jpg')
    elif (txnum.value == None) and platenum.value:
        if Tri=='R':
#            os.rename('raw\\'+prefix[0:(len(prefix)-len(str(platenum.value)))]+str(platenum.value)+'.jpg',name.value+'.jpg')
            shutil.copyfile('raw\\'+prefix[0:(len(prefix)-len(str(platenum.value)))]+str(platenum.value)+'.jpg','processed\\'+name.value+'.jpg')
            Tri = 'C'
        elif Tri=='C':
#            os.rename('raw\\'+prefix[0:(len(prefix)-len(str(platenum.value)))]+str(platenum.value)+'.jpg',name.value+'.jpg')
            shutil.copyfile('raw\\'+prefix[0:(len(prefix)-len(str(platenum.value)))]+str(platenum.value)+'.jpg','processed\\'+name.value+'.jpg')

            Tri = 'F'
        elif Tri=='F':
#            os.rename('raw\\'+prefix[0:(len(prefix)-len(str(platenum.value)))]+str(platenum.value)+'.jpg',name.value+'.jpg')
            shutil.copyfile('raw\\'+prefix[0:(len(prefix)-len(str(platenum.value)))]+str(platenum.value)+'.jpg','processed\\'+name.value+'.jpg')

            Tri = None

            
    
